from __future__ import annotations

import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook


REQUIRED_DIRS = ("00_scans", "01_text", "02_tickets", "03_final")
PROGRESS_COLUMNS = (
    "ticket_id",
    "scan_files",
    "ocr_status",
    "ticket_status",
    "check_notes",
    "final_status",
    "updated_at",
)
SUPPORTED_SCAN_SUFFIXES = {".jpg", ".jpeg", ".png", ".pdf"}
FINAL_STATUSES = {"raw", "ocr", "draft", "fix", "ready"}
TICKET_RE = re.compile(r"^ticket_(?P<ticket>\d{2})_page_(?P<page>\d{2})\.(?:jpg|jpeg|png|pdf)$", re.IGNORECASE)
RAW_TEXT_RE = re.compile(r"^ticket_(?P<ticket>\d{2})_raw\.md$")
REQUIRED_TICKET_SECTIONS = (
    "## Кратко",
    "## План ответа",
    "## Ответ на 5 минут",
    "## Что выучить точно",
    "## Вопросы для самопроверки",
    "## Проверить по скану",
)
OPTIONAL_EMPTY_TICKET_SECTIONS = {"## Проверить по скану"}
REQUIRED_TEXT_STUB_SECTIONS = (
    "## Source scans",
    "## OCR text",
    "## Manual checks",
    "## Notes",
)
OCR_PLACEHOLDER = "[paste OCR text here]"


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def ensure_root(root: Path) -> None:
    root.mkdir(parents=True, exist_ok=True)
    for name in REQUIRED_DIRS:
        (root / name).mkdir(exist_ok=True)


def progress_path(root: Path) -> Path:
    return root / "progress.xlsx"


def ensure_progress_workbook(root: Path) -> None:
    path = progress_path(root)
    if path.exists():
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "progress"
    sheet.append(PROGRESS_COLUMNS)
    workbook.save(path)


def load_rows(root: Path) -> dict[str, dict[str, str]]:
    ensure_progress_workbook(root)
    workbook = load_workbook(progress_path(root))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows: dict[str, dict[str, str]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value or "" for index, value in enumerate(values)}
        ticket_id = str(row.get("ticket_id", ""))
        if ticket_id:
            rows[ticket_id] = row
    workbook.close()
    return rows


def save_rows(root: Path, rows: dict[str, dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "progress"
    sheet.append(PROGRESS_COLUMNS)
    for ticket_id in sorted(rows):
        row = rows[ticket_id]
        final_status = row.get("final_status", "")
        if final_status and final_status not in FINAL_STATUSES:
            raise ValueError(f"invalid final_status for {ticket_id}: {final_status}")
        sheet.append([row.get(column, "") for column in PROGRESS_COLUMNS])
    workbook.save(progress_path(root))


def upsert_ticket(rows: dict[str, dict[str, str]], ticket_id: str) -> dict[str, str]:
    row = rows.setdefault(
        ticket_id,
        {
            "ticket_id": ticket_id,
            "scan_files": "",
            "ocr_status": "",
            "ticket_status": "",
            "check_notes": "",
            "final_status": "raw",
            "updated_at": "",
        },
    )
    row["ticket_id"] = ticket_id
    row["updated_at"] = now_iso()
    return row


def append_note(existing: str, note: str) -> str:
    parts = [part for part in existing.split("; ") if part]
    if note not in parts:
        parts.append(note)
    return "; ".join(parts)


def relative_scan_list(root: Path, files: list[Path]) -> str:
    return "\n".join(str(path.relative_to(root).as_posix()) for path in sorted(files))


def duplicate_names(paths: list[Path]) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for path in paths:
        name = path.name.lower()
        if name in seen:
            duplicates.add(path.name)
        seen.add(name)
    return sorted(duplicates)


def write_file(path: Path, content: str, force: bool = False) -> bool:
    if path.exists() and not force:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return True


def copy_file(source: Path, target: Path, force: bool = False) -> bool:
    if target.exists() and not force:
        return False
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, target)
    return True


def ticket_id_from_arg(value: str) -> str:
    digits = value.strip().removeprefix("ticket_")
    if not digits.isdigit():
        raise ValueError("--ticket must be a number like 01")
    return f"{int(digits):02d}"


def text_stub(ticket_id: str, scan_files: str) -> str:
    scans = [line for line in scan_files.splitlines() if line]
    scan_lines = "\n".join(f"- {line}" for line in scans) or "- "
    return f"""# Ticket {ticket_id} OCR Text

## Source scans
{scan_lines}

## OCR text

[paste OCR text here]

## Manual checks
- [ ] сверить текст по скану
- [ ] отметить неразборчивые места

## Notes
- [проверить по скану]
"""


def ticket_template(ticket_id: str) -> str:
    return f"""# Билет {ticket_id}. Название

## Кратко

## План ответа
1.
2.
3.

## Ответ на 5 минут

## Что выучить точно

## Вопросы для самопроверки
1.
2.
3.

## Проверить по скану
-
"""


def missing_sections(text: str) -> list[str]:
    return [section for section in REQUIRED_TICKET_SECTIONS if section not in text]


def meaningful_section_content(text: str, heading: str) -> str:
    content = section_content(text, heading)
    lines = []
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped or stripped in {"-", "1.", "2.", "3."}:
            continue
        lines.append(stripped)
    return "\n".join(lines).strip()


def validate_ticket_text(text: str) -> list[str]:
    issues = [f"missing header: {section.removeprefix('## ')}" for section in missing_sections(text)]
    if issues:
        return issues

    for section in REQUIRED_TICKET_SECTIONS:
        if section in OPTIONAL_EMPTY_TICKET_SECTIONS:
            continue
        if not meaningful_section_content(text, section):
            issues.append(f"empty section: {section.removeprefix('## ')}")

    check_content = meaningful_section_content(text, "## Проверить по скану")
    if check_content:
        issues.append("unresolved scan checks")
    return issues


def section_content(text: str, heading: str) -> str:
    start = text.find(heading)
    if start == -1:
        return ""
    start += len(heading)
    next_heading = text.find("\n## ", start)
    if next_heading == -1:
        return text[start:].strip()
    return text[start:next_heading].strip()


def missing_text_stub_sections(text: str) -> list[str]:
    return [section for section in REQUIRED_TEXT_STUB_SECTIONS if section not in text]


def has_non_empty_ocr_text(text: str) -> bool:
    content = section_content(text, "## OCR text")
    return bool(content and content != OCR_PLACEHOLDER)
